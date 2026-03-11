import re
import json
import pandas as pd
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl.styles import PatternFill, Alignment, Font
from src.process_dataframe import process_dataframe, sanitize_excel_values, normalize_text

class Splitter:
    def __init__(self):
        self.supported_projects = [
            "Be App",
            "Giao Hàng Nhanh",
            "Hafele",
            "Hanh Phuc Hospital/Hoan My Hospital",
            "ShopeeFood",
            "Tân Hiệp Phát",
            "HDBank",
            "PNJ",
            "HomeCredit",
            "HomeCredit (Cyber Fraud)"
        ]
        self.sheet_rules = self._load_sheet_rules()

    def _merge_interaction(self, df: pd.DataFrame, df_interaction: pd.DataFrame) -> pd.DataFrame:
        
        df.columns = df.columns.str.strip()
        df_interaction.columns = df_interaction.columns.str.strip()
        
        id_col = next((col for col in df.columns if 'id' == col.lower()), 'Id')
        post_id_col = next((col for col in df_interaction.columns if 'postid' == col.lower()), 'PostId')
        likes_col = next((c for c in df_interaction.columns if 'reaction' in c.lower()), None)
        shares_col = next((c for c in df_interaction.columns if 'share' in c.lower()), None)
        comments_col = next((c for c in df_interaction.columns if 'comment' in c.lower()), None)
        
        interaction_cols = [post_id_col]
        col_mapping = {}
        
        if likes_col:
            interaction_cols.append(likes_col)
            col_mapping[likes_col] = 'Likes'
        if shares_col:
            interaction_cols.append(shares_col)
            col_mapping[shares_col] = 'Shares'
        if comments_col:
            interaction_cols.append(comments_col)
            col_mapping[comments_col] = 'Comments'
        
        df_interaction_filtered = df_interaction[interaction_cols].copy()
        df_interaction_filtered = df_interaction_filtered.rename(columns=col_mapping)
        
        df_joined = df.merge(
            df_interaction_filtered,
            left_on=id_col,
            right_on=post_id_col,
            how='left'
        )

        if post_id_col in df_joined.columns:
            df_joined = df_joined.drop(columns=[post_id_col])
        
        for col in ['Likes', 'Shares', 'Comments']:
            if col in df_joined.columns:
                df_joined[col] = df_joined[col].fillna("not exist or close group")
        
        return df_joined
    
    def _filter_topic_type(self, df_raw: pd.DataFrame) -> pd.DataFrame:
        type_col = next((col for col in df_raw.columns if 'type' == col.lower()), "Type")
        types = ["fbPageTopic", "fbGroupTopic", "tiktokTopic", "fbUserTopic", "youtubeTopic"]
        types = [t.strip().lower() for t in types]
        df_filtered = df_raw[df_raw[type_col].astype(str).str.strip().str.lower().isin(types)]
        return df_filtered
    
    def _load_sheet_rules(self) -> Dict:
        try:
            rules_path = "rules/sheet_rules.json"
            print("Path: ", rules_path)
            with open(rules_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Warning: Could not load sheet_rules.json: {e}")
            return {}
    
    def get_supported_projects(self) -> List[str]:
        return self.supported_projects
    
    def _format_excel(self, df: pd.DataFrame, sheet_name: str = 'Data',
                     header_color: str = 'FF429EF5', 
                     column_mapping: Optional[Dict] = None, 
                     column_order: Optional[List] = None) -> BytesIO:
        """
        Format Excel với header màu tùy chỉnh và cột linh hoạt
        
        Args:
            df: DataFrame cần format
            sheet_name: Tên sheet (default: 'Data')
            header_color: Mã màu hex cho header (default: blue #429EF5)
            column_mapping: Dict để rename cột (optional)
            column_order: List thứ tự cột (optional)
        """
        df = df.copy()
        
        if column_mapping:
            df = df.rename(columns=column_mapping)
        
        if column_order:
            df = df[column_order]
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
            for cell in worksheet[1]:
                cell.fill = header_fill
        output.seek(0)
        return output
    
    def split_be_app(self, raw_file: BytesIO, interaction_file: BytesIO) -> BytesIO:
        df_raw = sanitize_excel_values(pd.read_excel(raw_file))
        df_interaction = sanitize_excel_values(pd.read_excel(interaction_file))
        df_filtered = self._filter_topic_type(df_raw)
        df_joined = self._merge_interaction(df_filtered, df_interaction)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_raw.to_excel(writer, sheet_name='Data', index=False)
            df_joined.to_excel(writer, sheet_name='Interaction', index=False)
        output.seek(0)
        return output
    
    def split_giao_hang_nhanh(self, raw_file: BytesIO, interaction_file: BytesIO) -> BytesIO:
        df = sanitize_excel_values(pd.read_excel(raw_file))
        df_interaction = sanitize_excel_values(pd.read_excel(interaction_file))
        
        labels_col = next((col for col in df.columns if 'labels1' == col.lower()), 'Labels1')
        df_buyer = df[df[labels_col].str.contains('Người mua hàng', case=False, na=False)].copy()
        df_receiver = df[df[labels_col].str.contains('Người nhận hàng', case=False, na=False)].copy()
        
        df_filtered = self._filter_topic_type(df)
        df_joined = self._merge_interaction(df_filtered, df_interaction)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)
            df_buyer.to_excel(writer, sheet_name='Người mua hàng', index=False)
            df_receiver.to_excel(writer, sheet_name='Người nhận hàng', index=False)
            df_joined.to_excel(writer, sheet_name='Interaction', index=False)
        output.seek(0)
        return output
    
    def split_hafele(self, file: BytesIO) -> BytesIO:
        df = sanitize_excel_values(pd.read_excel(file))
        category_col = next((col for col in df.columns if 'category' == col.lower()), 'Category')
        categories = df[category_col].dropna().unique()
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for category in categories:
                df_category = df[df[category_col] == category].copy()
                sheet_name = str(category)[:31]
                df_category.to_excel(writer, sheet_name=sheet_name, index=False)
        output.seek(0)
        return output
    
    def split_hospital(self, files: List[BytesIO], 
                       hp_interaction: BytesIO, 
                       hm_interaction: BytesIO) -> BytesIO:
        
        dfs = [pd.read_excel(file) for file in files]
        df = sanitize_excel_values(pd.concat(dfs, ignore_index=True))
        df_hp_interaction = sanitize_excel_values(pd.read_excel(hp_interaction))
        df_hm_interaction = sanitize_excel_values(pd.read_excel(hm_interaction))

        topic_col = next((col for col in df.columns if 'topic' == col.lower()), 'Topic')
        sentiment_col = next((col for col in df.columns if 'sentiment' == col.lower()), 'Sentiment')
        
        hospital_column_mapping = {
            "Labels1":	"System",
        }
        if hospital_column_mapping:
            df = df.rename(columns=hospital_column_mapping)

        columns = ["STT", "Title", "Content", "UrlComment", "UrlTopic", "Channel", "System", "PublishedDate", "Alert", "Time", "Type", "Likes", "Shares", "Comments"]
        
        def format_hospital_excel(df, columns, writer, sheet_name, sentiment=None):
            if len(df) > 0:
                df['STT'] = range(1, len(df) + 1)
                df = df.reindex(columns=columns).fillna("")
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                # Apply yellow header color
                worksheet = writer.sheets[sheet_name]
                header_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                for cell in worksheet[1]:
                    cell.fill = header_fill
            else:
                ws = writer.book.create_sheet(sheet_name)
                ws.merge_cells('A1:D2')
                cell = ws['A1']
                cell.value = 'KHÔNG CÓ NEGATIVE BUZZ' if sentiment == "Negative" else "KHÔNG CÓ NEUTRAL-POSITIVE BUZZ"
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                cell.font = Font(bold=True, size=14)
                
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_hp = df[df[topic_col].str.contains('Hanh Phuc Hospital', case=False, na=False)].copy()
            df_hp = self._merge_interaction(df_hp, df_hp_interaction)
            df_hp_neg = df_hp[df_hp[sentiment_col].str.contains('Negative', case=False, na=False)].copy()
            df_hp_pos = df_hp[~df_hp[sentiment_col].str.contains('Negative', case=False, na=False)].copy()
                
            df_hm = df[df[topic_col].str.contains('Hoan My Hospital', case=False, na=False)].copy()
            df_hm = self._merge_interaction(df_hm, df_hm_interaction)
            df_hm_neg = df_hm[df_hm[sentiment_col].str.contains('Negative', case=False, na=False)].copy()
            df_hm_pos = df_hm[~df_hm[sentiment_col].str.contains('Negative', case=False, na=False)].copy()
            
            format_hospital_excel(df_hp_neg, columns, writer, 'HP_Negative', "Negative")
            format_hospital_excel(df_hp_pos, columns, writer, 'HP_Neutral-Positive')

            format_hospital_excel(df_hm_neg, columns, writer, 'HM_Negative', "Negative")
            format_hospital_excel(df_hm_pos, columns, writer, 'HM_Neutral-Positive')
        
        return output
    
    def split_shopeefood(self, raw_file: BytesIO, demographic_file: BytesIO) -> BytesIO:
        df_raw = sanitize_excel_values(pd.read_excel(raw_file))
        df_demographic = sanitize_excel_values(pd.read_excel(demographic_file))
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_raw.to_excel(writer, sheet_name='Data', index=False)
            df_demographic.to_excel(writer, sheet_name='Demographic', index=False)
        output.seek(0)
        return output
    
    def split_tan_hiep_phat(self, file: BytesIO) -> BytesIO:
        df = sanitize_excel_values(pd.read_excel(file))
        label_col = next((col for col in df.columns if 'labels1' == col.lower()), 'Labels1')
        
        df_director = df[df[label_col].str.contains('BAN LÃNH ĐẠO', case=False, na=False)].copy()
        df_brand = df[df[label_col].str.contains('THƯƠNG HIỆU/CÔNG TY', case=False, na=False)].copy()
        
        mask_director = df[label_col].str.contains('BAN LÃNH ĐẠO', case=False, na=False)
        mask_brand = df[label_col].str.contains('THƯƠNG HIỆU/CÔNG TY', case=False, na=False)
        df_product = df[~(mask_director | mask_brand)].copy()
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_director.to_excel(writer, sheet_name='Ban Lãnh Đạo', index=False)
            df_brand.to_excel(writer, sheet_name='Thương Hiệu', index=False)
            df_product.to_excel(writer, sheet_name='Sản phẩm', index=False)
        output.seek(0)
        return output
    
    def split_pnj(self, raw_file: BytesIO, demographic_file: BytesIO) -> BytesIO:

        df_original = sanitize_excel_values(pd.read_excel(raw_file))
        df_demographic = sanitize_excel_values(pd.read_excel(demographic_file))

        print("Total:", len(df_original))

        df_google = df_original[df_original["SiteName"].str.contains("google.com", case=False, na=False)]

        df = df_original[~df_original["SiteName"].str.contains("google.com", case=False, na=False)].copy()

        df = process_dataframe(df)

        pnj_rules = self.sheet_rules.get("PNJ", {})

        if not pnj_rules:
            raise ValueError("Không tìm thấy rules cho PNJ trong sheet_rules.json")

        # normalize text 1 lần
        df["Text_norm"] = df["Text"].fillna("").astype(str).str.lower().apply(normalize_text)

        df["SheetName"] = "Khác"

        # apply rules bằng vectorization
        for rule in pnj_rules:

            sheet_name = rule["sheet"]

            keywords = [
                normalize_text(k.lower())
                for k in rule["keywords"]
            ]

            # escape regex
            keywords = [re.escape(k) for k in keywords]

            pattern = "|".join(keywords)

            mask = df["Text_norm"].str.contains(pattern, na=False, regex=True)

            df.loc[mask & (df["SheetName"] == "Khác"), "SheetName"] = sheet_name

        df = df.drop(columns=["Text", "Text_norm"])

        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            df_original.to_excel(writer, sheet_name="Data", index=False)
            df_google.to_excel(writer, sheet_name="Google", index=False)
            df_demographic.to_excel(writer, sheet_name="Demographic", index=False)

            rule_sheets = [rule["sheet"] for rule in pnj_rules]

            for sheet_name in rule_sheets:

                df_sheet = df[df["SheetName"] == sheet_name].copy()

                if not df_sheet.empty:
                    df_sheet = df_sheet.drop(columns=["SheetName"])
                else:
                    df_sheet = df.drop(columns=["SheetName"]).iloc[0:0]

                safe_sheet_name = str(sheet_name)[:31]

                df_sheet.to_excel(
                    writer,
                    sheet_name=safe_sheet_name,
                    index=False
                )

        output.seek(0)

        return output


    def split_hdbank(self, file: BytesIO) -> Tuple[BytesIO, BytesIO]:
        df = sanitize_excel_values(pd.read_excel(file))
        topic_col = next((col for col in df.columns if 'topic' == col.lower()), 'Topic')
        
        hdbank_topics = ['HDBank', 'HD Saison', 'HD Securities', 'Sovico Group']
        competitor_topics = ['Techcombank', 'MBBank', 'VPBank', 'ACB Bank']
        
        df_hdbank = df[df[topic_col].str.strip().isin(hdbank_topics)].copy()
        df_competitors = df[df[topic_col].str.strip().isin(competitor_topics)].copy()
        
        output_hdbank = BytesIO()
        with pd.ExcelWriter(output_hdbank, engine='openpyxl') as writer:
            for topic in hdbank_topics:
                df_topic = df_hdbank[df_hdbank[topic_col] == topic].copy()
                if len(df_topic) > 0:
                    df_topic.to_excel(writer, sheet_name=topic, index=False)
        output_hdbank.seek(0)
        
        output_competitors = BytesIO()
        with pd.ExcelWriter(output_competitors, engine='openpyxl') as writer:
            df_competitors.to_excel(writer, sheet_name='Đối thủ', index=False)
        output_competitors.seek(0)
        
        return output_hdbank, output_competitors
    
    def split_home(self, raw_file: BytesIO, interaction_file: BytesIO) -> BytesIO:
        df = sanitize_excel_values(pd.read_excel(raw_file))
        df_interaction = sanitize_excel_values(pd.read_excel(interaction_file))

        topic_col = next((col for col in df.columns if 'topic' == col.lower()), 'Topic')
        bnpl_sheet = ['Kredivo', 'Fundiin', 'Lotte Paylater', 'Muadee']
        home_sheet = ['Home Credit']
        finance_sheet = ["F88", "FE Credit", "HD Saison", "MCredit", "Mirae Asset", "Shinhan Finance"]
        
        topic_series = df[topic_col].astype(str).str.strip()
        df_bnpl = df[topic_series.isin(bnpl_sheet)].copy()
        df_home = df[topic_series.isin(home_sheet)].copy()
        df_finance = df[topic_series.isin(finance_sheet)].copy()
 
        df_filtered = self._filter_topic_type(df)
        df_joined = self._merge_interaction(df_filtered, df_interaction)
    
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_bnpl.to_excel(writer, sheet_name='BNPL', index=False)
            df_home.to_excel(writer, sheet_name='Home Credit', index=False)
            df_finance.to_excel(writer, sheet_name='Financial', index=False)
            df_joined.to_excel(writer, sheet_name='Interaction', index=False)
        output.seek(0)
        return output
    
    def format_home_processed(self, raw_file: BytesIO, interaction_file: BytesIO) -> BytesIO:
        
        df = sanitize_excel_values(pd.read_excel(raw_file))
        df_interaction = sanitize_excel_values(pd.read_excel(interaction_file))
        df_joined = self._merge_interaction(df, df_interaction)
        
        df_joined["Type (comment or post)"] = df_joined["Type"].astype(str).str.lower().apply(
            lambda x: "Comment" if "comment" in x else "Post"
        )
        
        column_mapping = {
            "Topic": "Group",
            "UrlComment": "URL",
            "PublishedDate": "Date",
            "SiteName": "Source"
        }
        
        df_joined["No."] = range(1, len(df_joined) + 1)
        
        column_order = ["No.", "Group", "Content", "Type (comment or post)", "URL",
                       "Date", "Source", "Channel", "Author", "Likes", "Shares", "Comments"]
        
        return self._format_excel(df_joined, sheet_name='Data', header_color='FF429EF5', 
                                 column_mapping=column_mapping, 
                                 column_order=column_order)

    def split(self, project: str, output_filename: Optional[str] = None, **kwargs) -> Dict:
        try:
            if project == "Be App":
                output = self.split_be_app(kwargs['raw_file'], kwargs['interaction_file'])
                filename = f"{output_filename}.xlsx" if output_filename else 'Kompa x Be App.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "Giao Hàng Nhanh":
                output = self.split_giao_hang_nhanh(kwargs['raw_file'], kwargs['interaction_file'])
                filename = f"{output_filename}.xlsx" if output_filename else 'Kompa x Giao Hàng Nhanh.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "Hafele":
                output = self.split_hafele(kwargs['file'])
                filename = f"{output_filename}.xlsx" if output_filename else 'Kompa x Hafele.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "Hanh Phuc Hospital/Hoan My Hospital":
                output = self.split_hospital(kwargs['files'], kwargs['hp_interaction'], kwargs['hm_interaction'])
                filename = f"{output_filename}.xlsx" if output_filename else f'Kompa x {project}.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "ShopeeFood":
                output = self.split_shopeefood(kwargs['raw_file'], kwargs['demographic_file'])
                filename = f"{output_filename}.xlsx" if output_filename else 'Kompa x ShopeeFood.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "Tân Hiệp Phát":
                output = self.split_tan_hiep_phat(kwargs['file'])
                filename = f"{output_filename}.xlsx" if output_filename else 'Kompa x Tân Hiệp Phát.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "HDBank":
                output_hdbank, output_competitors = self.split_hdbank(kwargs['file'])
                if output_filename:
                    filename1 = f"{output_filename}.xlsx"
                    filename2 = f"{output_filename}_Competitors.xlsx"
                else:
                    filename1 = 'Kompa x HDBank.xlsx'
                    filename2 = 'Kompa x HDBank_Competitors.xlsx'
                return {'files': [output_hdbank, output_competitors], 'filenames': [filename1, filename2], 
                       'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "PNJ":
                output = self.split_pnj(kwargs['raw_file'], kwargs['demographic_file'])
                filename = f"{output_filename}.xlsx" if output_filename else 'Kompa x PNJ.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "HomeCredit":
                output = self.split_home(kwargs['raw_file'], kwargs['interaction_file'])
                filename = f"{output_filename}.xlsx" if output_filename else 'Kompa x HomeCredit.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            elif project == "HomeCredit (Cyber Fraud)":
                output = self.format_home_processed(kwargs['raw_file'], kwargs['interaction_file'])
                filename = f"{output_filename}.xlsx" if output_filename else 'Kompa_Daily_Cyber_Fraud.xlsx'
                return {'files': [output], 'filenames': [filename], 'success': True, 'message': 'Đã tách thành công'}
            
            else:
                return {'files': [], 'filenames': [], 'success': False, 
                       'message': f'Project "{project}" chưa được hỗ trợ'}
        
        except Exception as e:
            return {'files': [], 'filenames': [], 'success': False, 'message': f'Lỗi: {str(e)}'}
